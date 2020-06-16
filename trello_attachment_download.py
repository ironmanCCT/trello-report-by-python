#trello attachment bulk downloader v0.1 by mwong@convergencect.com
## 1. export your file to xlsx format with 2 coloumns, [casename, attachments] using TrelloExport Google Chrome Add-on. 
## 2. place this script and xlsx export file into same folder
## 3. requires python 3.7+ and the imported packages.
## 4. you can adjust the sleep timer to be shorter than the default .1 seconds to download files faster. But be careful to watch out for remote host blocking due to too many get requests.
## 5. Added Support For Threading (1 thread per card)
import os
import requests
from openpyxl import load_workbook
import re
import pprint
import time
import threading
import datetime
import sys
import math

config = {
            'workbook': 'TrelloExport_20200610180734.xlsx',
            'worksheet': 'TrelloExport',
            'numberOfThreads': 8,   #testing on 16 GB Ram, Runtime: 1:19:10.365137
                                    #(dont set number of threads > cards in your TrelloExport
                                    #, It's possible if the thread count is too high and it starts downloading, but has to wait to wr
            'zFillNumber':  4       # number of digits required for padding file numbers e.g. 0001 for up to 9999 documents
                                    # documents will be named according to their position in the TrelloExport Attachment Column Cell.
         }

begin_time = datetime.datetime.now()
print('being time: ' + str(begin_time))
workbook = load_workbook(filename=config['workbook'])
worksheet = workbook[config['worksheet']]
cwd = os.getcwd()  ##find the current working directory to save files to
print('current working directory: ' + cwd)
zFillNumber = config["zFillNumber"]
def trelloBulkDownloadAttachments(start, end):
    for i in range(start,end):
        print(' starting loop number:' + str(i))
        if(not worksheet.cell(i,2).value):
            print(str(i)+' has no attachments')
        else:
            attachment = (worksheet.cell(i,2).value)
            results = []
            results.append(re.findall('(\[.*\]) (\(.*\)) (https.*)$', attachment, flags= re.MULTILINE))  #this will parse out multiple attachments in one cell to an array of tuples
            caseName = worksheet.cell(i,1).value
            caseName = caseName.replace('/', '-')
            path = cwd + '\\' + caseName
            #optional print commands to see variables
            #print('worksheet.cell('+str(i)+').value: ')
            #print(worksheet.cell(i,2).value)
            #print('results: ' + str(results))
            #print('caseName: ' + caseName)
            #caseNames.append(worksheet.cell(i,1).value)
            #print('path exists? ' + str(os.path.isdir(path)))
            #print("%s is the staged path" % path)
            if not os.path.isdir(path):
                try:
                    os.mkdir(path)
                except OSError:
                    print ("Creation of the directory %s failed" % path)
                else:
                    print ("Successfully created the directory %s " % path)
            else:
                print ("%s directory probably exisits " % caseName)
                continue
            
            #download the attachments and save them to the caseName Directory
            for result in results:
                #print(result)
                try:
                    for index, name_and_id_and_url in enumerate(result):
                        filename = name_and_id_and_url[0].replace('[','').replace(']','')  #get the name from the first item in tuple
                        fileNumber = str(index+1).zfill(zFillNumber)
                        if('https:' in filename):
                            print('attachment is not a file')
                            continue
                        elif filename == '':
                            filename =  fileNumber +'-fileId-'+str(name_and_id_and_url[1])
                            saveToDirectory = path +'\\'+ filename
                            url = name_and_id_and_url[2] #get the url from the third item in tuple
                            r = requests.get(url, allow_redirects=True)
                            open(saveToDirectory, 'wb').write(r.content)
                        else:
                            #time.sleep()
                            saveToDirectory = path + '\\' + fileNumber + '-' + filename
                            url = name_and_id_and_url[2] #get the url from the third item in tuple
                            r = requests.get(url, allow_redirects=True)
                            open(saveToDirectory, 'wb').write(r.content)
                            #print("successfully downloaded: " + saveToDirectory)
                except OSError:
                    print ("unable to download file: %s" % saveToDirectory)
                    print('ERROR: ' + str(sys.exc_info()))
                else:
                    print ("Successfully downloaded all files to  the directory %s " % path)
                    continue

#split downloads into x amount of threads set in configuration
numberOfThreads = config["numberOfThreads"]          
batchSize = math.floor(worksheet.max_row/numberOfThreads) #if max_rows = 101 then batchSize is 25 with remainder 1 row.
modWorksheetMaxRow = worksheet.max_row % numberOfThreads #need list of 1, 25, 50, 75, 100, 101
rangeList = []
#calculate the row range for each thread based on thread size. 
for i in range(0,numberOfThreads):
    if i < numberOfThreads-1:
        rangeList.append((batchSize*i+1,batchSize*i + batchSize + 1))
    else:
        rangeList.append((batchSize*i+1,batchSize*i + batchSize + modWorksheetMaxRow + 1))
#initialize the threads with the function for downloading attachments by card column                          
thread_list = []
for (x,y) in rangeList:
        thread = threading.Thread(target=trelloBulkDownloadAttachments, args=(x, y))
        thread_list.append(thread) 
#start the threads, join the threads
for thread in thread_list:
    thread.start()
for thread in thread_list:
    thread.join()
#print the runtime at finish. 
print('\n Runtime:' + str(datetime.datetime.now() - begin_time))
